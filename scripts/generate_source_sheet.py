"""
Generates the source submission spreadsheet.
Output: docs/source_submission_template.xlsx
Run: py scripts/generate_source_sheet.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ──────────────────────────────────────────────────────────────────
DARK       = "1E1E28"
ACCENT     = "2D6A4F"
ACCENT_LT  = "EDF7F1"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F7"
MID_GRAY   = "D0D0D5"
GRAY_TEXT  = "9090A0"
INACTIVE   = "EBEBED"
INACTIVE_T = "C0C0C8"
SILVER     = "B4B4BC"

def F(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def font(bold=False, size=10, color=DARK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)

thin  = Side(style="thin",   color=MID_GRAY)
thick = Side(style="medium", color=DARK)
green = Side(style="medium", color=ACCENT)

def border(*sides):
    kw = dict(left=thin, right=thin, top=thin, bottom=thin)
    kw.update(sides)
    return Border(**kw)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(indent=1, wrap=False):
    return Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=wrap)

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sources"

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 48
ws.column_dimensions["D"].width = 40
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 22

# ── Row 1: Title banner ───────────────────────────────────────────────────────
ws.row_dimensions[1].height = 38
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value     = "Market Intelligence  |  Source Submission"
c.font      = font(bold=True, size=15, color=WHITE)
c.fill      = F(DARK)
c.alignment = left(indent=2)

# ── Row 2: Country / Market label ────────────────────────────────────────────
ws.row_dimensions[2].height = 16
ws.merge_cells("A2:C2")
ws.merge_cells("D2:F2")

ws["A2"].value     = "COUNTRY / MARKET"
ws["A2"].font      = font(bold=True, size=8, color=GRAY_TEXT)
ws["A2"].alignment = left(indent=1)

ws["D2"].value     = "SUBMITTED BY"
ws["D2"].font      = font(bold=True, size=8, color=GRAY_TEXT)
ws["D2"].alignment = left(indent=1)

# ── Row 3: Country / Market input | Submitted By input ───────────────────────
ws.row_dimensions[3].height = 26
ws.merge_cells("A3:C3")
ws.merge_cells("D3:F3")

ws["A3"].fill      = F(LIGHT_GRAY)
ws["A3"].border    = Border(left=green, right=thin, top=thin, bottom=thin)
ws["A3"].alignment = left(indent=2)
ws["A3"].font      = font(size=11)

ws["D3"].fill      = F(LIGHT_GRAY)
ws["D3"].border    = Border(left=green, right=thin, top=thin, bottom=thin)
ws["D3"].alignment = left(indent=2)
ws["D3"].font      = font(size=11)

# ── Row 4: Number of Sources label ───────────────────────────────────────────
ws.row_dimensions[4].height = 16
ws.merge_cells("A4:F4")

ws["A4"].value     = "NUMBER OF SOURCES  (enter a number — up to 100 rows will activate)"
ws["A4"].font      = font(bold=True, size=8, color=GRAY_TEXT)
ws["A4"].alignment = left(indent=1)

# ── Row 5: Number of Sources input (key cell = A5) ───────────────────────────
ws.row_dimensions[5].height = 34
ws.merge_cells("A5:F5")

ws["A5"].fill      = F(ACCENT_LT)
ws["A5"].border    = Border(left=green, right=thin, top=thin, bottom=thin)
ws["A5"].alignment = left(indent=2)
ws["A5"].font      = font(bold=True, size=18)

# ── Row 6: Column headers ─────────────────────────────────────────────────────
ws.row_dimensions[6].height = 20
headers = [
    "#",
    "Source Name  *",
    "Source URL  *",
    "Description  *",
    "Relationship Type",
    "Business Domain",
]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=6, column=col)
    c.value     = h
    c.font      = font(bold=True, size=9, color=WHITE)
    c.fill      = F(ACCENT)
    c.alignment = center()
    c.border    = border()

# ── Rows 7-106: Data rows ─────────────────────────────────────────────────────
for row in range(7, 107):
    ws.row_dimensions[row].height = 20
    idx = row - 6  # source number (1-based)
    bg  = LIGHT_GRAY if idx % 2 == 0 else WHITE

    for col in range(1, 7):
        c = ws.cell(row=row, column=col)
        c.fill   = F(bg)
        c.border = border()
        c.font   = font(size=9)
        c.alignment = left(wrap=(col == 4))

    # Row number column
    num = ws.cell(row=row, column=1)
    num.value     = idx
    num.alignment = center()
    num.font      = font(size=8, color=GRAY_TEXT)

# ── Data validation: dropdowns ────────────────────────────────────────────────
rel_options = (
    '"Government,Association,Customer,Partner,Competitor,General News"'
)
dom_options = (
    '"BER - Built Environment,'
    'EDU - Education,'
    'MFG - Manufacturing,'
    'HLS - Healthcare,'
    'RCC - Retail & Commerce,'
    'CTE - Culture & Tourism,'
    'PSS - Public Sector"'
)

dv_rel = DataValidation(
    type="list", formula1=rel_options,
    allow_blank=True, showDropDown=False,
    showErrorMessage=True,
    errorTitle="Invalid entry",
    error='Choose from the list, or leave blank if unsure.'
)
dv_dom = DataValidation(
    type="list", formula1=dom_options,
    allow_blank=True, showDropDown=False,
)
ws.add_data_validation(dv_rel)
ws.add_data_validation(dv_dom)
dv_rel.sqref = "E7:E106"
dv_dom.sqref = "F7:F106"

# ── Conditional formatting: gray out rows beyond A5 ──────────────────────────
#    Formula applied from row 7: row index = ROW()-6
#    Gray if: row index > entered number  (or > 3 if A5 is blank)
gray_fill = F(INACTIVE)
gray_font = Font(color=INACTIVE_T, size=9)

ws.conditional_formatting.add(
    "A7:F106",
    FormulaRule(
        formula=["ROW()-6>IF(ISBLANK($A$5),3,$A$5)"],
        fill=gray_fill,
        font=gray_font,
    ),
)

# ── Freeze header rows ────────────────────────────────────────────────────────
ws.freeze_panes = "A7"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "docs/source_submission_template.xlsx"
wb.save(out)
print(f"Saved: {out}")
